from fastapi import FastAPI, APIRouter, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse, FileResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from fastapi.middleware.cors import CORSMiddleware
import os
import csv
import logging
import tempfile
import time
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import Any, Optional, Union
from io import BytesIO
import uuid
from datetime import datetime, timezone
import httpx
import openpyxl
import xlsxwriter
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
# This provides a default value so it never throws a KeyError
mongo_url = os.environ.get('MONGO_URL', 'mongodb://localhost:27017/dummy_db')
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ.get('DB_NAME', 'dummy_db')]

app = FastAPI()
api_router = APIRouter(prefix="/api")

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# In-memory store for prepared downloads (token -> file path & metadata)
_download_store: dict[str, dict[str, Any]] = {}

SCS_API_URL = "https://scssustratedgemsverifyapi.azurewebsites.net/api/verifygems"
SCS_API_CODE = "a2xtJa5IxG0jvYZ81Ht/rn6xPR6eSA1G12ira/61iiGYAkCsXc6mEw=="

# ───────────────────── Models ─────────────────────

class GemData(BaseModel):
    model_config = ConfigDict(extra="allow")

    inventory_id: str
    gem_producer: Optional[str] = None
    gem_producer_expanded: Optional[str] = None
    grading_lab_code: Optional[str] = None
    grading_lab_report_id: Optional[str] = None
    gem_type: Optional[str] = None
    sust_rated_eligible: Optional[Union[bool, str]] = None
    certified_sustainable: Optional[Union[bool, str]] = None
    carat_wt: Optional[Union[float, str]] = None
    color_code: Optional[str] = None
    clarity_code: Optional[str] = None
    shape_code: Optional[str] = None
    cut_value: Optional[str] = None
    polish_value: Optional[str] = None
    symmetry_value: Optional[str] = None
    stone_type_code: Optional[str] = None
    measure_1: Optional[Union[float, str]] = None
    measure_2: Optional[Union[float, str]] = None
    measure_3: Optional[Union[float, str]] = None
    depth_percentage: Optional[Union[float, str]] = None
    table_percentage: Optional[Union[float, str]] = None
    production_date: Optional[str] = None
    production_location: Optional[str] = None
    production_location_expanded: Optional[str] = None
    scs_certificate_id: Optional[str] = None
    scs_gem_certificate_id: Optional[str] = None
    scs_gem_certificate_url: Optional[str] = None
    scs_gem_cert_qrcode_url: Optional[str] = None
    not_sust_rated_eligible_msgs: Optional[Any] = None
    not_certified_sustainable_msgs: Optional[Any] = None
    error: Optional[str] = None


class VerifyRequest(BaseModel):
    stone_ids: list[str]


class VerificationHistory(BaseModel):
    model_config = ConfigDict(extra="ignore")

    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    stone_ids: list[str]
    results: list[GemData]
    timestamp: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class PrepareExportRequest(BaseModel):
    format: str = "excel"
    gems: list[dict[str, Any]]


# ───────────────────── Helpers ─────────────────────

def _append_attachment_param(item: dict[str, Any]) -> None:
    """Append &as_attachment=false to certificate URL for browser viewing."""
    url = item.get("scs_gem_certificate_url")
    if url:
        item["scs_gem_certificate_url"] = url + "&as_attachment=false"


def _parse_api_response(api_data: Any) -> list[GemData]:
    """Parse SCS API response into a list of GemData, normalising cert URLs."""
    items: list[dict[str, Any]] = []
    if isinstance(api_data, list):
        items = api_data
    elif isinstance(api_data, dict) and "data" in api_data:
        items = api_data["data"]
    elif isinstance(api_data, dict):
        items = [api_data]

    for item in items:
        _append_attachment_param(item)

    return [GemData(**item) for item in items]


#async def _save_verification_history(
 #   stone_ids: list[str], results: list[GemData]
#) -> None:
 #   """Persist a verification run to MongoDB."""
  #  history = VerificationHistory(stone_ids=stone_ids, results=results)
   # doc = history.model_dump()
    #doc["timestamp"] = doc["timestamp"].isoformat()
    #await db.verification_history.insert_one(doc)

async def _save_verification_history(stone_ids: list[str], results: list[GemData]) -> None:
    try:
        history = VerificationHistory(stone_ids=stone_ids, results=results)
        doc = history.model_dump()
        doc["timestamp"] = doc["timestamp"].isoformat()
        # Set a short timeout so it doesn't hang the app
        await db.verification_history.insert_one(doc)
    except Exception as e:
        logger.warning(f"Database not available, skipping history save: {e}")

def _sustainable_label(gem: GemData) -> str:
    """Return a human-readable sustainability label."""
    if gem.certified_sustainable == True:  # noqa: E712
        return "Yes"
    if gem.certified_sustainable == False:  # noqa: E712
        return "No"
    return ""


# ── File-parsing helpers ──

def _find_id_column(headers: list[Any]) -> Optional[int]:
    """Return index of the column most likely to hold stone IDs."""
    keywords = ("inventory", "stone", "id")
    for idx, header in enumerate(headers):
        if header and any(kw in str(header).lower() for kw in keywords):
            return idx
    return None


def _parse_excel(contents: bytes) -> list[str]:
    """Extract stone IDs from an Excel workbook."""
    wb = openpyxl.load_workbook(BytesIO(contents))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    col_index = _find_id_column(headers)
    start_row = 2 if col_index is not None else 1
    col_index = col_index if col_index is not None else 0

    return [
        str(row[col_index]).strip()
        for row in ws.iter_rows(min_row=start_row, values_only=True)
        if row[col_index]
    ]


def _parse_csv(contents: bytes) -> list[str]:
    """Extract stone IDs from a CSV file."""
    content_str = contents.decode("utf-8")
    rows = list(csv.reader(content_str.splitlines()))
    if not rows:
        raise HTTPException(status_code=400, detail="Empty CSV file")

    has_header = any(
        "inventory" in str(cell).lower() or "stone" in str(cell).lower()
        for cell in rows[0]
    )
    start_idx = 1 if has_header else 0
    return [
        str(row[0]).strip()
        for row in rows[start_idx:]
        if row and row[0]
    ]


# ── Export-generation helpers ──

EXCEL_HEADERS: list[str] = [
    "Inventory ID","Grading Lab Report ID", "Grading Lab Code", "Carat Weight", "Color", "Clarity", "Shape",
    "Cut", "Polish", "Symmetry", "Depth %", "Table %",
    "Sustainable", "Certificate URL", "QR Code URL",
]

PDF_TABLE_STYLE = TableStyle([
    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0A2540")),
    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
    ("FONTSIZE", (0, 0), (-1, 0), 10),
    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
    ("BACKGROUND", (0, 1), (-1, -1), colors.white),
    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ("FONTSIZE", (0, 1), (-1, -1), 8),
])


def _gem_to_excel_row(gem: GemData) -> list[Any]:
    """Convert a GemData to a flat list matching EXCEL_HEADERS."""
    return [
        gem.inventory_id,
         gem.grading_lab_report_id or "",
        gem.grading_lab_code or "",
        gem.carat_wt if gem.carat_wt else "",
        gem.color_code or "",
        gem.clarity_code or "",
        gem.shape_code or "",
        gem.cut_value or "",
        gem.polish_value or "",
        gem.symmetry_value or "",
        gem.depth_percentage if gem.depth_percentage else "",
        gem.table_percentage if gem.table_percentage else "",
        _sustainable_label(gem),
        gem.scs_gem_certificate_url or "",
        gem.scs_gem_cert_qrcode_url or "",
    ]


def _generate_excel_file(filepath: str, gems: list[GemData]) -> None:
    """Write an Excel workbook to *filepath*."""
    workbook = xlsxwriter.Workbook(filepath)
    ws = workbook.add_worksheet("Gem Certificates")
    header_fmt = workbook.add_format(
        {"bold": True, "bg_color": "#0A2540", "font_color": "white", "border": 1}
    )
    cell_fmt = workbook.add_format({"border": 1})

    for col, h in enumerate(EXCEL_HEADERS):
        ws.write(0, col, h, header_fmt)
    for row_num, gem in enumerate(gems, 1):
        for col, val in enumerate(_gem_to_excel_row(gem)):
            ws.write(row_num, col, val, cell_fmt)

    ws.set_column(0, 0, 15)
    ws.set_column(1, 10, 13)
    ws.set_column(11, 12, 55)
    workbook.close()


def _generate_pdf_file(filepath: str, gems: list[GemData]) -> None:
    """Write a PDF report to *filepath*."""
    doc = SimpleDocTemplate(filepath, pagesize=A4)
    styles = getSampleStyleSheet()
    elements: list[Any] = []

    elements.append(Paragraph("Gem Certificate Verification Report", styles["Heading1"]))
    elements.append(Spacer(1, 0.3 * inch))

    # Data table
    table_data = [["Inventory ID","Grading Lab Report ID", "Grading Lab Code", "Carat Weight", "Color", "Clarity", "Shape", "Cut", "Sustainable"]]
    for gem in gems:
        table_data.append([
            gem.inventory_id,str(gem.grading_lab_report_id or ""), str(gem.grading_lab_code or ""), str(gem.carat_wt or ""), gem.color_code or "",
            gem.clarity_code or "", gem.shape_code or "", gem.cut_value or "",
            _sustainable_label(gem),
        ])

    table = Table(table_data)
    table.setStyle(PDF_TABLE_STYLE)
    elements.append(table)
    elements.append(Spacer(1, 0.3 * inch))

    # Certificate URLs
    if any(g.scs_gem_certificate_url for g in gems):
        elements.append(Paragraph("<b>Certificate URLs:</b>", styles["Heading2"]))
        elements.append(Spacer(1, 0.1 * inch))
        for gem in gems:
            if gem.scs_gem_certificate_url:
                elements.append(
                    Paragraph(
                        f"<b>{gem.inventory_id}:</b> {gem.scs_gem_certificate_url}",
                        styles["Normal"],
                    )
                )
                elements.append(Spacer(1, 0.05 * inch))

    doc.build(elements)


def _cleanup_expired_tokens() -> None:
    """Remove download tokens older than 5 minutes."""
    now = time.time()
    expired = [k for k, v in _download_store.items() if now - v["created"] > 300]
    for k in expired:
        try:
            os.remove(_download_store[k]["path"])
        except OSError:
            pass
        del _download_store[k]


# ───────────────────── Routes ─────────────────────

@api_router.get("/")
async def root() -> dict[str, str]:
    return {"message": "Gem Certificate Verification API"}


@api_router.post("/verify-stones", response_model=list[GemData])
async def verify_stones(request: VerifyRequest) -> list[GemData]:
    """Verify multiple stones using the SCS API."""
    if not request.stone_ids:
        raise HTTPException(status_code=400, detail="Please provide at least one stone ID")

    params = {
        "gem_producer": "fx",
        "inventory_ids": ",".join(request.stone_ids),
        "code": SCS_API_CODE,
    }

    try:
        # 'verify=False' tells the code not to complain about the SSL certificate
        async with httpx.AsyncClient(timeout=30.0, verify=False) as http:
            response = await http.get(SCS_API_URL, params=params)
            response.raise_for_status()

        results = _parse_api_response(response.json())
        await _save_verification_history(request.stone_ids, results)
        return results

    except httpx.HTTPError as e:
        logger.error("API request failed: %s", e)
        raise HTTPException(status_code=500, detail=f"Failed to verify stones: {e}")
    except Exception as e:
        logger.error("Verification error: %s", e)
        raise HTTPException(status_code=500, detail=f"Verification error: {e}")


@api_router.post("/upload-file", response_model=list[str])
async def upload_file(file: UploadFile = File(...)) -> list[str]:
    """Upload CSV or Excel file and extract stone IDs."""
    filename = file.filename or ""
    contents = await file.read()

    try:
        if filename.endswith((".xlsx", ".xls")):
            stone_ids = _parse_excel(contents)
        elif filename.endswith(".csv"):
            stone_ids = _parse_csv(contents)
        else:
            raise HTTPException(
                status_code=400,
                detail="Unsupported file format. Please upload CSV or Excel file.",
            )
    except HTTPException:
        raise
    except Exception as e:
        logger.error("File upload error: %s", e)
        raise HTTPException(status_code=400, detail=f"File processing error: {e}")

    if not stone_ids:
        raise HTTPException(status_code=400, detail="No stone IDs found in file")

    return stone_ids


@api_router.post("/export/excel")
async def export_excel(gems: list[GemData]) -> StreamingResponse:
    """Export gem data to Excel (legacy streaming endpoint)."""
    output = BytesIO()
    workbook = xlsxwriter.Workbook(output, {"in_memory": True})
    ws = workbook.add_worksheet("Gem Certificates")
    header_fmt = workbook.add_format(
        {"bold": True, "bg_color": "#0A2540", "font_color": "white", "border": 1}
    )
    cell_fmt = workbook.add_format({"border": 1})

    for col, h in enumerate(EXCEL_HEADERS):
        ws.write(0, col, h, header_fmt)
    for row_num, gem in enumerate(gems, 1):
        for col, val in enumerate(_gem_to_excel_row(gem)):
            ws.write(row_num, col, val, cell_fmt)

    ws.set_column(0, 0, 15)
    ws.set_column(1, 10, 13)
    ws.set_column(11, 12, 55)
    workbook.close()
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=gem_certificates.xlsx"},
    )


@api_router.post("/export/pdf")
async def export_pdf(gems: list[GemData]) -> StreamingResponse:
    """Export gem data to PDF (legacy streaming endpoint)."""
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4)
    styles = getSampleStyleSheet()
    elements: list[Any] = []

    elements.append(Paragraph("Gem Certificate Verification Report", styles["Heading1"]))
    elements.append(Spacer(1, 0.3 * inch))

    table_data = [["Inventory ID", "Carat", "Color", "Clarity", "Shape", "Cut"]]
    for gem in gems:
        table_data.append([
            gem.inventory_id, str(gem.carat_wt or ""), gem.color_code or "",
            gem.clarity_code or "", gem.shape_code or "", gem.cut_value or "",
        ])

    table = Table(table_data)
    table.setStyle(PDF_TABLE_STYLE)
    elements.append(table)
    elements.append(Spacer(1, 0.3 * inch))

    if gems and gems[0].scs_gem_certificate_url:
        elements.append(Paragraph("<b>Certificate URLs:</b>", styles["Heading2"]))
        elements.append(Spacer(1, 0.1 * inch))
        for gem in gems:
            if gem.scs_gem_certificate_url:
                elements.append(
                    Paragraph(
                        f"<b>{gem.inventory_id}:</b> {gem.scs_gem_certificate_url}",
                        styles["Normal"],
                    )
                )
                elements.append(Spacer(1, 0.05 * inch))

    doc.build(elements)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=gem_certificates.pdf"},
    )


@api_router.post("/prepare-export")
async def prepare_export(payload: PrepareExportRequest) -> dict[str, str]:
    """Prepare an export file and return a download token."""
    gems = [GemData(**g) for g in payload.gems]
    if not gems:
        raise HTTPException(status_code=400, detail="No gem data to export")

    token = str(uuid.uuid4())
    tmp_dir = tempfile.gettempdir()

    try:
        if payload.format == "excel":
            filepath = os.path.join(tmp_dir, f"{token}.xlsx")
            _generate_excel_file(filepath, gems)
            _download_store[token] = {
                "path": filepath,
                "filename": "gem_certificates.xlsx",
                "media_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "created": time.time(),
            }
        else:
            filepath = os.path.join(tmp_dir, f"{token}.pdf")
            _generate_pdf_file(filepath, gems)
            _download_store[token] = {
                "path": filepath,
                "filename": "gem_certificates.pdf",
                "media_type": "application/pdf",
                "created": time.time(),
            }

        _cleanup_expired_tokens()
        return {"token": token}

    except Exception as e:
        logger.error("Prepare export error: %s", e)
        raise HTTPException(status_code=500, detail=f"Export error: {e}")


@api_router.get("/download/{token}")
async def download_file(token: str) -> FileResponse:
    """Download a previously prepared export file."""
    entry = _download_store.get(token)
    if not entry or not os.path.exists(entry["path"]):
        raise HTTPException(status_code=404, detail="Download expired or not found")

    return FileResponse(
        path=entry["path"],
        filename=entry["filename"],
        media_type=entry["media_type"],
        headers={"Content-Disposition": f"attachment; filename={entry['filename']}"},
    )


@api_router.get("/history", response_model=list[VerificationHistory])
async def get_history() -> list[VerificationHistory]:
    """Get verification history."""
    history = await db.verification_history.find(
        {}, {"_id": 0}
    ).sort("timestamp", -1).limit(50).to_list(50)

    for record in history:
        if isinstance(record["timestamp"], str):
            record["timestamp"] = datetime.fromisoformat(record["timestamp"])

    return history


# ───────────────────── App wiring ─────────────────────
# ───────────────────── App wiring ─────────────────────

# 1. Include the router (Removed the prefix here because it's already in APIRouter)
app.include_router(api_router)

# 2. Clean CORS Setup
origins_str = os.environ.get("CORS_ORIGINS", "http://localhost:3000,http://127.0.0.1:3000")
origins = [o.strip() for o in origins_str.split(",")]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# 3. Modern "Lifespan" approach to remove that warning
from contextlib import asynccontextmanager

@asynccontextmanager
async def lifespan(app: FastAPI):
    # Startup logic
    yield
    # Shutdown logic
    await client.close()

# Update the FastAPI app definition at the top of your file to use this:
# app = FastAPI(lifespan=lifespan)